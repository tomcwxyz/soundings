"""Sport England — Active Lives Adult Survey client.

Downloads the "Tables 1-5 Levels of Activity" Excel workbook from
Sport England's S3 bucket and reads the LA-level sheet (Table 3).

The workbook is published annually; the URL changes each release.
The default URL points to the Nov 2024-25 publication (Apr 2026).
Override via the `url` constructor param when a new version is published.

Open Government Licence v3.0.
"""

import io
from collections.abc import Iterator
from typing import Any

import httpx
import openpyxl

# Sport England S3 bucket — URL changes per publication. This is the
# Nov 2024-25 release (published April 2026).
ACTIVE_LIVES_URL = (
    "https://sportengland-production-files.s3.eu-west-2.amazonaws.com/s3fs-public/"
    "2026-04/Active%20Lives%20Adult%20Survey%20report%20Nov%2024-25%20"
    "Tables%201-5%20Levels%20of%20activity.xlsx"
)

# Sheet name for LA-level activity rates.
LA_SHEET = "Table 3 Levels Local Authorit"

# Data starts at row 8 (0-indexed) after title + multi-row headers.
HEADER_ROWS = 8


class SportEnglandActiveLivesClient:
    """Download the Active Lives Excel workbook and parse Table 3."""

    def __init__(
        self,
        http_client: httpx.AsyncClient | None = None,
        url: str = ACTIVE_LIVES_URL,
    ) -> None:
        self._client = http_client
        self._owns_client = http_client is None
        self._url = url

    async def fetch_workbook(self) -> bytes:
        client = self._client or httpx.AsyncClient(timeout=120.0)
        try:
            response = await client.get(self._url, follow_redirects=True)
            response.raise_for_status()
            return response.content
        finally:
            if self._owns_client:
                await client.aclose()

    def read_la_sheet(self, content: bytes) -> Iterator[dict[str, Any]]:
        """Yield one dict per LA data row from Table 3.

        The sheet has a multi-row header (rows 0-7), then data rows
        grouped by region. Each row has either a county-level GSS code
        in col 0 or a district-level GSS code in col 2. We read both
        and key by column index.

        Returns dicts with keys:
            gss_code: str — the 9-character GSS code
            la_name: str
            active_rate: float | None — % active (150+ min/week), Nov 2024-25
            fairly_active_rate: float | None — % fairly active (30-149 min)
            inactive_rate: float | None — % inactive (<30 min)
        """
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # Sheet name may be truncated by Excel's 31-char limit
        if LA_SHEET in wb.sheetnames:
            ws = wb[LA_SHEET]
        else:
            # Try a prefix match — Excel truncates long sheet names to 31 chars
            candidates = [s for s in wb.sheetnames if s.startswith("Table 3")]
            if not candidates:
                raise ValueError(f"Sheet '{LA_SHEET}' not found. Available: {wb.sheetnames}")
            ws = wb[candidates[0]]

        rows = list(ws.iter_rows(values_only=True))

        for row in rows[HEADER_ROWS:]:
            if row is None:
                continue
            # Try col 2 first (district/unitary level — our LTLAs)
            gss_code = _clean_str(row[2]) if len(row) > 2 else None
            la_name = _clean_str(row[3]) if len(row) > 3 else None

            # Fall back to col 0 (county level)
            if not gss_code or not _is_gss_code(gss_code):
                gss_code = _clean_str(row[0]) if len(row) > 0 else None
                la_name = _clean_str(row[1]) if len(row) > 1 else None

            if not gss_code or not _is_gss_code(gss_code):
                continue

            # Nov 2024-25 columns (latest period):
            # col 33 = Active population total, col 34 = Active rate
            # col 37 = Fairly Active population total, col 38 = Fairly Active rate
            # col 41 = Inactive population total, col 42 = Inactive rate
            active_rate = _coerce_float(row[34]) if len(row) > 34 else None
            fairly_active_rate = _coerce_float(row[38]) if len(row) > 38 else None
            inactive_rate = _coerce_float(row[42]) if len(row) > 42 else None

            yield {
                "gss_code": gss_code,
                "la_name": la_name or "",
                "active_rate": active_rate,
                "fairly_active_rate": fairly_active_rate,
                "inactive_rate": inactive_rate,
            }


def _clean_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _is_gss_code(s: str) -> bool:
    """Check if a string looks like a GSS code (E + 8 digits, W + 8 digits)."""
    if not s or len(s) != 9:
        return False
    return s[0] in ("E", "W", "S", "N") and s[1:].isdigit()


def _coerce_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
    except (ValueError, TypeError):
        return None
    import math

    if not math.isfinite(f):
        return None
    return f
